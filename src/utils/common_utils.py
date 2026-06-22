import re
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import random
import os, requests
from PIL import Image
import hashlib
from datetime import datetime, timezone
from dateutil.relativedelta import relativedelta
import json
from typing import List, Dict 


def load_contacted_users(contacted_users_file):
    if not os.path.exists(contacted_users_file):
        os.makedirs(os.path.dirname(contacted_users_file), exist_ok=True)
        return set()
    with open(contacted_users_file, 'r', encoding='utf-8') as f:
        return {line.strip() for line in f if line.strip()}


def parse_cookie_string(cookie_string: str) -> List[Dict]:
    """解析从浏览器复制的Cookie字符串"""
    try:
        # 方法1：如果是JSON格式
        if cookie_string.strip().startswith('['):
            cookies = json.loads(cookie_string)
            print(f"已加载JSON格式Cookie，共 {len(cookies)} 个")
            return cookies
        
        # 方法2：如果是键值对格式 (key=value; key2=value2)
        cookies = []
        pairs = cookie_string.split(';')
        for pair in pairs:
            if '=' in pair:
                key, value = pair.strip().split('=', 1)
                cookies.append({
                    'name': key.strip(),
                    'value': value.strip(),
                    'domain': '.amazon.com',
                    'path': '/',
                    'secure': True,
                    'httpOnly': False
                })
        #print(f"已解析Cookie字符串，共 {len(cookies)} 个")
        return cookies
    except Exception as e:
        print(f"解析Cookie时出错: {e}")
        return []


def parse_product_dimensions(dim_str):
    """
    解析亚马逊商品长宽高字符串，提取长、宽、高数值和单位
    支持格式：1x1x1cm / 1 x 1 x 1 cm / 2.5 X 3.8 x 4.2 inch / 1,000 x 2,500.5 x 3,000 cm
    
    :param dim_str: 输入的尺寸字符串（可包含其他无关文本，如"尺寸：15.5x20.8x8 cm"）
    :return: 元组 (length, width, height, unit)（数值为float，单位小写），匹配失败返回None
    """
    # 步骤1：预处理 - 移除数值中的千分位逗号（如1,000 → 1000）
    dim_str_clean = re.sub(r'(\d),(\d)', r'\1\2', dim_str)
    
    # 步骤2：正则匹配核心模式（支持x/X、任意空格、整数/小数、字母单位）
    # 分组说明：\1=长, \2=宽, \3=高, \4=单位
    pattern = r'(\d+\.?\d*)\s*[Xx]\s*(\d+\.?\d*)\s*[Xx]\s*(\d+\.?\d*)\s*([a-zA-Z]+)'
    match = re.search(pattern, dim_str_clean)  # search支持字符串中包含其他文本
    
    default_value = [0.0, 0.0, 0.0, '']
    # 步骤3：匹配失败直接返回None
    if not match:
        return default_value
    
    # 步骤4：提取分组并转换数值
    length_str, width_str, height_str, unit = match.groups()
    try:
        length = float(length_str)
        width = float(width_str)
        height = float(height_str)
    except ValueError:
        # 数值转换失败（正则已过滤非数值，此情况极少）
        return default_value
    
    # 步骤5：单位统一转小写，方便后续处理（如单位转换）
    unit = unit.lower()
    
    return (length, width, height, unit)


def get_text_response_ds(context, prompt, model_name="deepseek-v4-flash", api_key=""):
    try:
        TEXT_CLIENT = OpenAI(api_key=api_key, base_url="https://api.deepseek.com/v1")
        title_response = TEXT_CLIENT.chat.completions.create(
            model=model_name,
            messages=[
                {"role": "system", "content": context},
                {"role": "user", "content": prompt},
            ],
            stream=False
        )

        response_content = title_response.choices[0].message.content #.replace('\n', '<br>')
        return response_content
    except Exception as e:
        print(f'Error: {e}')
        return ''


def insert_validation_column(excel_path, sheet_name, insert_indexs, column_names, options=["-1,0,3"], err_msg=["只能选择-1/0/3"]):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    # 
    if options is not None:
        for idx, col_name, opt, e_msg in zip(insert_indexs, column_names, options, err_msg):
            ws.insert_cols(idx)

            # 字母
            col_letter = get_column_letter(idx)

            # 表头
            ws[f"{col_letter}1"] = col_name

            # 创建下拉单选
            dv = DataValidation(
                type="list",
                formula1=f'{opt}',
                allow_blank=True,
                showErrorMessage=True,
                error=e_msg
            )

            ws.add_data_validation(dv)

            # 应用到整列（从第 2 行开始）
            dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")
    else:
        for idx, col_name in zip(insert_indexs, column_names):
            ws.insert_cols(idx)

            # 字母
            col_letter = get_column_letter(idx)

            # 表头
            ws[f"{col_letter}1"] = col_name

    wb.save(excel_path)


def random_11_digits():
    return ''.join(str(random.randint(0, 9)) for _ in range(11))


def calc_upc_check_digit(upc11: str) -> str:
    """
    计算 UPC-A 校验位
    upc11: 11位字符串
    """
    assert len(upc11) == 11 and upc11.isdigit()

    odd_sum = sum(int(upc11[i]) for i in range(0, 11, 2))
    even_sum = sum(int(upc11[i]) for i in range(1, 11, 2))

    total = odd_sum * 3 + even_sum
    check_digit = (10 - (total % 10)) % 10
    return str(check_digit)


def generate_upc_list(count: int, mode="random", output_file_path="", start: int = 0, write_out=False):
    """
    批量生成 UPC-A
    start: 起始11位数字（int）
    count: 生成数量
    mode: 随机生成还是按顺序从start生成
    """
    used_upc_list = [] if not os.path.exists(output_file_path) else [x.strip() for x in open(output_file_path).readlines()]

    generated_upc_list = []

    offset = 0
    for _ in range(count):
        while 1:
            digit_string = random_11_digits() if mode=='random' else start + offset
            upc11 = str(digit_string).zfill(11)
            check_digit = calc_upc_check_digit(upc11)
            upc12 = upc11 + check_digit

            if upc12 not in used_upc_list:
                generated_upc_list.append(upc12)
                break

            offset += 1

    if write_out:
        with open(output_file_path, 'a') as fd:
            fd.writelines('\n'.join(generated_upc_list))
            fd.writelines('\n')


    return generated_upc_list

def calc_ean13_check_digit(ean12: str) -> int:
    """
    计算 EAN-13 校验位
    """
    if len(ean12) != 12 or not ean12.isdigit():
        raise ValueError("EAN 主体必须是 12 位数字")

    total = 0
    for i, ch in enumerate(ean12):
        digit = int(ch)
        # EAN-13：从左数，偶数位 *3，奇数位 *1
        total += digit * (3 if (i + 1) % 2 == 0 else 1)

    return (10 - (total % 10)) % 10


def generate_ean13_list(count: int, output_file_path="", prefix: str = "") -> str:
    """
    生成一批合法的 EAN-13
    prefix：可选前缀（如国家码）

    国家/地区代码：
        690-699：中国
        000-019/030-039/060-139：美国、加拿大
        400-440：德国
        500-509：英国
        750-759：墨西哥
        779：阿根廷
        789-790：巴西
    """
    used_ean_list = [] if not os.path.exists(output_file_path) else [x.strip() for x in open(output_file_path).readlines()]
    generated_ean_list = []
    
    if prefix and not prefix.isdigit():
        raise ValueError("prefix 必须是数字")

    body_length = 12 - len(prefix)
    if body_length <= 0:
        raise ValueError("prefix 长度不能超过 12")

    offset = 0
    for _ in range(count):
        while 1:
            body = prefix + ''.join(str(random.randint(0, 9)) for _ in range(body_length))
            check_digit = calc_ean13_check_digit(body)
            ean13 = body + str(check_digit)

            if ean13 not in used_ean_list:
                generated_ean_list.append(ean13)
                break
            
            offset += 1

    with open(output_file_path, 'a') as fd:
        fd.writelines('\n'.join(generated_ean_list))
        fd.writelines('\n')

    return generated_ean_list


def human_scroll(page, steps=5):
    """模拟人类滚动"""
    for i in range(steps):
        page.mouse.wheel(0, random.randint(300, 800))
        page.wait_for_timeout(random.randint(500, 1200))


def generate_hash_digits(input_str, algorithm='md5', to_pure_digits=False):
    """
    对输入字符串生成哈希值（可选转为纯数字串）
    
    参数：
        input_str: 待处理的原始字符串
        algorithm: 哈希算法，可选 'md5'/'sha1'/'sha256' 等
        to_pure_digits: 是否转为纯数字串（True/False）
    
    返回：
        哈希字符串（或纯数字串）
    """
    try:
        # 1. 检查算法是否支持
        if algorithm not in hashlib.algorithms_available:
            raise ValueError(f"不支持的哈希算法：{algorithm}")
        
        # 2. 将字符串编码为字节（必须步骤）
        input_bytes = input_str.encode('utf-8')
        
        # 3. 创建哈希对象并计算哈希值
        hash_obj = hashlib.new(algorithm)
        hash_obj.update(input_bytes)
        # 获取十六进制哈希串（如 MD5 是 32 位，SHA256 是 64 位）
        hash_hex = hash_obj.hexdigest()
        
        # 4. 可选：转为纯数字串（十六进制转十进制）
        if to_pure_digits:
            # int(hash_hex, 16) 将十六进制字符串转为十进制整数
            hash_digits = str(int(hash_hex, 16))
            return hash_digits
        else:
            return hash_hex
    
    except Exception as e:
        print(f"生成哈希值失败：{e}")
        return None


def generate_utc_time_strings():
    """
    生成仅精确到秒、无毫秒的当前UTC时间和一年后UTC时间
    格式：YYYY-MM-DDTHH:MM:SS+00:00
    :return: (当前时间字符串, 一年后时间字符串)
    """
    # 1. 获取当前UTC时间（带时区）
    current_utc = datetime.now(timezone.utc)
    # 2. 计算一年后的UTC时间（自动处理闰年/月份天数）
    one_year_later_utc = current_utc + relativedelta(years=1, days=-1)
    
    # 3. 格式化：仅保留到秒，去掉毫秒，严格匹配目标格式
    # %Y=年 %m=月 %d=日 T=分隔符 %H=时 %M=分 %S=秒 +00:00=UTC时区
    time_format = "%Y-%m-%dT%H:%M:%S+00:00"
    current_time_str = current_utc.strftime(time_format)
    one_year_later_str = one_year_later_utc.strftime(time_format)
    
    return current_time_str, one_year_later_str


def get_adspower_ws(user_id, api_key="4188a4ee49461bef870df28cefc9ecef008bdc717c5b3d88", base_url="http://127.0.0.1:50325"):
    """获取 AdsPower 的 CDP 调试地址"""
    headers = {"api-key": api_key}
    start_url = f"{base_url}/api/v1/browser/start?user_id={user_id}"
    
    try:
        resp = requests.get(start_url, headers=headers).json()
        if resp["code"] == 0:
            return resp["data"]["ws"]["puppeteer"]
        else:
            print(f"❌ AdsPower 启动失败: {resp['msg']}")
            return None
    except Exception as e:
        print(f"❌ 请求 AdsPower API 出错: {e}")
        return None
